from flask import Flask, redirect, url_for, render_template, request, session
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from datetime import datetime
import smtplib
from email.message import EmailMessage


app = Flask(__name__)

app.secret_key = "vrcs0715"

@app.route("/", methods=["POST", "GET"])
def home():
	if request.method == "POST":
		vet = request.form["nm"]
		session["vet"] = vet
		return redirect(url_for("feedlot"))
	else:
		return render_template("Home.html")

@app.route("/feedlot", methods=["POST", "GET"])
def feedlot():
    global vet
    if "vet" in session:
        vet = session["vet"]
    if request.method == "POST":
            feedlots = request.form["nm"]
            session["feedlot"] = feedlots
            list = ["J. Trent Fox, PhD, DVM", ["NM", "TX", "OK", "CO", "ID", "AZ", "NV", "NE", "WA", "CA", "KS"], "Travis McCarty, DVM", ["KS", "OK", "CO", "NE"], "Del Miles, DVM, MS", ["TX", "KS", "OK", "CO", "NE", "CA", "ID", "WA"], "Tom Portillo, DVM", ["TX", "NM", "KS"], "Randy Hunter, DVM", ["WY", "CO", "NE", "MT", "KS"], "Bob Smith, DVM, MS", ["TX", "KS", "OK", "CO", "NE"], "Miles Theurer, DVM, PhD", ["KS"], "Ryan McCollum, DVM", ["TX", "KS", "OK", "NM"], "Simon Timmermans, DVM, MS", ["IA", "WI", "MO", "ND", "MN", "NE", "KS", "ID", "WA", "MI"], "Jim Sears, DVM", "John Lynch, DVM"]
            count = 0
            for i in list:
                    if i != vet:
                        count += 1
                        continue
                    else:
                        count += 1
                        if feedlots[-8:-6] in list[count]:
                            return redirect(url_for("NotDairy"))
                        else:
                            return redirect(url_for("home"))
            #return redirect(url_for("home"))
    else:
          return render_template("Feedlot.html")





@app.route("/NotDairy", methods=["POST", "GET"])
def NotDairy():
    if request.method == "POST":
            drug = request.form["Drugs"]
            quantity = request.form["Quantity"]
            size = request.form["Sizes"]
            species = request.form["species"]
            session["drug"] = drug
            session["quantity"] = quantity
            session["size"] = size
            session["species"] = species
            drug1 = request.form["Drugs1"]
            quantity1 = request.form["Quantity1"]
            size1 = request.form["Sizes1"]
            species1 = request.form["species1"]
            session["drug1"] = drug1
            session["quantity1"] = quantity1
            session["size1"] = size1
            session["species1"] = species1
            drug2 = request.form["Drugs2"]
            quantity2 = request.form["Quantity2"]
            size2 = request.form["Sizes2"]
            species2 = request.form["species2"]
            session["drug2"] = drug2
            session["quantity2"] = quantity2
            session["size2"] = size2
            session["species2"] = species2
            drug3 = request.form["Drugs3"]
            quantity3 = request.form["Quantity3"]
            size3 = request.form["Sizes3"]
            species3 = request.form["species3"]
            session["drug3"] = drug3
            session["quantity3"] = quantity3
            session["size3"] = size3
            session["species3"] = species3
            drug4 = request.form["Drugs4"]
            quantity4 = request.form["Quantity4"]
            size4 = request.form["Sizes4"]
            species4 = request.form["species4"]
            session["drug4"] = drug4
            session["quantity4"] = quantity4
            session["size4"] = size4
            session["species4"] = species4
            drug5 = request.form["Drugs5"]
            quantity5 = request.form["Quantity5"]
            size5 = request.form["Sizes5"]
            species5 = request.form["species5"]
            session["drug5"] = drug5
            session["quantity5"] = quantity5
            session["size5"] = size5
            session["species5"] = species5

            drug6 = request.form["Drugs6"]
            quantity6 = request.form["Quantity6"]
            size6 = request.form["Sizes6"]
            species6 = request.form["species6"]
            session["drug6"] = drug6
            session["quantity6"] = quantity6
            session["size6"] = size6
            session["species6"] = species6
            drug7 = request.form["Drugs7"]
            quantity7 = request.form["Quantity7"]
            size7 = request.form["Sizes7"]
            species7 = request.form["species7"]
            session["drug7"] = drug7
            session["quantity7"] = quantity7
            session["size7"] = size7
            session["species7"] = species7
            drug8 = request.form["Drugs8"]
            quantity8 = request.form["Quantity8"]
            size8 = request.form["Sizes8"]
            species8 = request.form["species8"]
            session["drug8"] = drug8
            session["quantity8"] = quantity8
            session["size8"] = size8
            session["species8"] = species8
            drug9 = request.form["Drugs9"]
            quantity9 = request.form["Quantity9"]
            size9 = request.form["Sizes9"]
            species9 = request.form["species9"]
            session["drug9"] = drug9
            session["quantity9"] = quantity9
            session["size9"] = size9
            session["species9"] = species9
            drug10 = request.form["Drugs10"]
            quantity10 = request.form["Quantity10"]
            size10 = request.form["Sizes10"]
            species10 = request.form["species10"]
            session["quantity10"] = quantity10
            session["size10"] = size10
            session["species10"] = species10
            drug11 = request.form["Drugs11"]
            quantity11 = request.form["Quantity11"]
            size11 = request.form["Sizes11"]
            species11 = request.form["species11"]
            session["quantity11"] = quantity11
            session["size11"] = size11
            session["species11"] = species11
            drug12 = request.form["Drugs12"]
            quantity12 = request.form["Quantity12"]
            size12 = request.form["Sizes12"]
            species12 = request.form["species12"]
            session["quantity12"] = quantity12
            session["size12"] = size12
            session["species12"] = species12
            drug13 = request.form["Drugs13"]
            quantity13 = request.form["Quantity13"]
            size13 = request.form["Sizes13"]
            species13 = request.form["species13"]
            session["quantity13"] = quantity13
            session["size13"] = size13
            session["species13"] = species13
            drug14 = request.form["Drugs14"]
            quantity14 = request.form["Quantity14"]
            size14 = request.form["Sizes14"]
            species14 = request.form["species14"]
            session["quantity14"] = quantity14
            session["size14"] = size14
            session["species14"] = species14
            drug15 = request.form["Drugs15"]
            quantity15 = request.form["Quantity15"]
            size15 = request.form["Sizes15"]
            species15 = request.form["species15"]
            session["quantity15"] = quantity15
            session["size15"] = size15
            session["species15"] = species15
            drug16 = request.form["Drugs16"]
            quantity16 = request.form["Quantity16"]
            size16 = request.form["Sizes16"]
            species16 = request.form["species16"]
            session["quantity16"] = quantity16
            session["size16"] = size16
            session["species16"] = species16
            drug17 = request.form["Drugs17"]
            quantity17 = request.form["Quantity17"]
            size17 = request.form["Sizes17"]
            species17 = request.form["species17"]
            session["quantity17"] = quantity17
            session["size17"] = size17
            session["species17"] = species17
            drug18 = request.form["Drugs18"]
            quantity18 = request.form["Quantity18"]
            size18 = request.form["Sizes18"]
            species18 = request.form["species18"]
            session["quantity18"] = quantity18
            session["size18"] = size18
            session["species18"] = species18
            drug19 = request.form["Drugs19"]
            quantity19 = request.form["Quantity19"]
            size19 = request.form["Sizes19"]
            species19 = request.form["species19"]
            session["quantity19"] = quantity19
            session["size19"] = size19
            session["species19"] = species19
            drug20 = request.form["Drugs20"]
            quantity20 = request.form["Quantity20"]
            size20 = request.form["Sizes20"]
            species20 = request.form["species20"]
            session["quantity20"] = quantity20
            session["size20"] = size20
            session["species20"] = species20
            session["drug20"] = drug20
            session["drug19"] = drug19
            session["drug18"] = drug18
            session["drug17"] = drug17
            session["drug16"] = drug16
            session["drug15"] = drug15
            session["drug14"] = drug14
            session["drug13"] = drug13
            session["drug12"] = drug12
            session["drug11"] = drug11
            session["drug10"] = drug10


            return redirect(url_for("info"))
    else:
        return render_template("NotDairy.html")

@app.route("/info", methods=["POST", "GET"])
def info():
    if request.method == "POST":
        productdict = {
            "Acepromazine" : "Acepromazine maleate, 10 mg/ml",
            "Advocin" : "Danofloxacin Mesylate, 180mg/ml",
            "Anased La 100mg/ml" : "Xylazine, 100 mg/ml",
            "Banamine " : "Flunixin Meglumine, 50 mg/ml",
            "Banamine Transdermal" : "Flunixin meglumine, 50 mg/ml",
            "Banamine Paste" : "Flunixin Meglumine Paste, 1500 mg/30-g syringe",
            "Baytril 100" : "Enrofloxacin, 100 mg/ml",
            "BioMycin 200" : "Oxytetracycline, 200 mg/ml",
            "Caldex CMPK" : "10.8g-8g-2.5g-1.6g-75g/500 ml",
            "Ceftiflex" : "Ceftiofur, 50 mg/ml",
            "Cefenil RTU" : "Ceftiofur, 50 mg/ml",
            "Cystorelin" : "Gonadorelin, 50 mcg/ml",
            "Dormosedan" : "Detomadine HCl, 10 mg/ml",
            "Dormosedan Gel" : "",
            "Depomedrol" : "Methylprednisolone, 10 mg/ml",
            "Dexamethasone" : "Dexamethasone, 2 mg/ml",
            "Dormosedan" : "10 mg/ml",
            "Dormosedan Gel" : "7.6 mg/mL detomidine hydrochloride",
            "Draxxin" : "Tulathromycin, 100 mg/ml",
            "E-SE" : "Sodium selenite, 5.48 mg/ml; Vitamin E, 68 IU/ml",
            "Equi Bute Paste Apple" : "Phenylbutazone, 20gm/60ml",
            "Epinephrine" : "Epinephrine, 1 mg/ml",
            "EstroPlan" : "Cloprostenol, 250 mcg/ml",
            "Estrumate" : "Cloprostenol, 250 mcg/ml",
            "Excede" : "Ceftiofur Crystalline Free acid, 200 mg/ml",
            "Excenel RTU" : "Ceftiofur HCl, 50 mg/ml",
            "Factrel" : "Gonadorelin, 50 mg/ml",
            "Fertagyl" : "Gonadorelin, 43 mcg/ml",
            "Iodine" : "",
            "Lasix" : "Furosemide, 5%",
            "Lidocaine" : "Lidocaine HCl, 2.0%",
            "Lutalyse" : "Dinoprost promethamine, 5 mg/ml",
            "Lutalyse (high con)" : "Dinoprost promethamine, 12.5 mg/ml",
            "Micotil" : "Tilmicosin, 300mg/ml",
            "Multimin 90" : "Zn 60mg; Mn 10mg; Se 5 mg;Cu 15 mg /ml",
            "Mu-SE" : "5 mg Selenium, 50 mg Vit E/ml",
            "Naxcel" : "Ceftiofur, 50 mg/ml",
            "NeoPoly Dex Ophthalmic Ointment" : "Neomycin, Polymyxin B, Dexamethasone, 3.5 g",
            "Norfenicol" : "Florfenicol, 300 mg/ml",
            "NuFlor" : "Florfenicol, 300 mg/ml",
            "Oxytocin" : "Oxytocin, 20 USP/ml",
            "Polyflex" : "Ampicillin, 25 g",
            "Prostamate" : "Dinoprost tromethamine, 5 mg/ml",
            "Phenylzone Paste" : "Phenylbutazone, 1g/3ml ",
            "Rabvac 3" : "Killed Rabies Virus Vaccine",
            "ResFlor Gold" : "Florfenicol, 300 mg/ml; Flunixin meglumine 16.5 mg/ml",
            "Salix or Furosemide" : "Furosemide, 50 mg/ml",
            "Salmonella Newport SRP" : "",
            "Sodium Bicarbonate " : "Sodium Bicarbonate, 8.4%",
            "Sodium iodide" : "Sodium Iodide, 200 mg/ml",
            "Sterile Water" : "Water, 100%",
            "Super B complex" : "Thiamine HCL 100mg, Riboflavin 5mg, Pyridoxine HCL 10mg, Niacinamide 100mg, d-Panthenol 10mg, Cyanocobalamin 100mcg per ml",
            "Sustain III boluses" : "Sulfamethazine, 8.02 g/bolus",
            "Synchsure" : "Cloprostenol, 250 mcg/ml",
            "Thiamine 200" : "Thiamine HCl, 200 mg/ml",
            "Thiamine 500" : "Thiamine HCl, 500 mg/ml",
            "Uniprim" : "67 mg trimethoprim, 333 mg sulfadiazine/g",
            "Vetribute" : "Phenylbutazone, 20 gm",
            "Vitamin B-12, 1000mcg" : "Cyanocobalamin, 1000mcg/ml",
            "Vitamin B Complex" : "Thiamine, 50 mg; Riboflavin, 2mg; Pyridoxine, 2mg; Niacin, 100mg; Panthenol, 10mg; Cobalt, 0.4 ppm/ml",
            "Vitamin E IU" : "Tocopherol, 300 mg/ml or 500 mg/ml",
            "Vitamin K" : "Phytonadione, 10 mg/ml",
            "Vitamin C" : "Sodium ascorbate, 250 mg/ml",
            "Zactran" : "Gamithromycin, 150 mg/ml",
            "Zuprevo" : "Tildopirosin, 180 mg/ml"
        }

        product1 = ""
        product2 = ""
        product3 = ""
        product4 = ""
        product5 = ""
        product6 = ""
        product7 = ""
        product8 = ""
        product9 = ""
        product10 = ""
        product11 = ""
        product12 = ""
        product13 = ""
        product14 = ""
        product15 = ""
        product16 = ""
        product17 = ""
        product18 = ""
        product19 = ""
        product20 = ""
        product21 = ""
        product22 = ""

        if "drug" in session:
            drug1 = session["drug"]
            quantity1 = session["quantity"]
            size1 = session["size"]
            feedlot = session["feedlot"]
            vet = session["vet"]
            species1 = session["species"]
            for x in productdict.keys():
                if x == drug1:
                    product1 = productdict.get(x)
                    break
                else:
                    continue
        if "drug1" in session:
            drug2 = session["drug1"]
            quantity2 = session["quantity1"]
            size2 = session["size1"]
            species2 = session["species1"]
            for x in productdict.keys():
                if x == drug2:
                    product2 = productdict.get(x)
                    break
                else:
                    continue
        if "drug2" in session:
            drug3 = session["drug2"]
            quantity3 = session["quantity2"]
            size3 = session["size2"]
            species3 = session["species2"]
            for x in productdict.keys():
                if x == drug3:
                    product3 = productdict.get(x)
                    break
                else:
                    continue
        if "drug3" in session:
            drug4 = session["drug3"]
            quantity4 = session["quantity3"]
            size4 = session["size3"]
            species4 = session["species3"]
            for x in productdict.keys():
                if x == drug4:
                    product4 = productdict.get(x)
                    break
                else:
                    continue
        if "drug4" in session:
            drug5 = session["drug4"]
            quantity5 = session["quantity4"]
            size5 = session["size4"]
            species5 = session["species4"]
            for x in productdict.keys():
                if x == drug5:
                    product5 = productdict.get(x)
                    break
                else:
                    continue
        if "drug5" in session:
            drug6 = session["drug5"]
            quantity6 = session["quantity5"]
            size6 = session["size5"]
            species6 = session["species5"]
            for x in productdict.keys():
                if x == drug6:
                    product6 = productdict.get(x)
                    break
                else:
                    continue
        if "drug6" in session:
            drug7 = session["drug6"]
            quantity7 = session["quantity6"]
            size7 = session["size6"]
            species7 = session["species6"]
            for x in productdict.keys():
                if x == drug7:
                    product7 = productdict.get(x)
                    break
                else:
                    continue
        if "drug7" in session:
            drug8 = session["drug7"]
            quantity8 = session["quantity7"]
            size8 = session["size7"]
            species8 = session["species7"]
            for x in productdict.keys():
                if x == drug8:
                    product8 = productdict.get(x)
                    break
                else:
                    continue
        if "drug8" in session:
            drug9 = session["drug8"]
            quantity9 = session["quantity8"]
            size9 = session["size8"]
            species9 = session["species8"]
            for x in productdict.keys():
                if x == drug9:
                    product9 = productdict.get(x)
                    break
                else:
                    continue
        if "drug9" in session:
            drug10 = session["drug9"]
            quantity10 = session["quantity9"]
            size10 = session["size9"]
            species10 = session["species9"]
            for x in productdict.keys():
                if x == drug10:
                    product10 = productdict.get(x)
                    break
                else:
                    continue
        if "drug10" in session:
            drug11 = session["drug10"]
            quantity11 = session["quantity10"]
            size11 = session["size10"]
            species11 = session["species10"]
            for x in productdict.keys():
                if x == drug11:
                    product11 = productdict.get(x)
                    break
                else:
                    continue
        if "drug11" in session:
            drug12 = session["drug11"]
            quantity12 = session["quantity11"]
            size12 = session["size11"]
            species12 = session["species11"]
            for x in productdict.keys():
                if x == drug12:
                    product12 = productdict.get(x)
                    break
                else:
                    continue
        if "drug12" in session:
            drug13 = session["drug12"]
            quantity13 = session["quantity12"]
            size13 = session["size12"]
            species13 = session["species12"]
            for x in productdict.keys():
                if x == drug13:
                    product13 = productdict.get(x)
                    break
                else:
                    continue
        if "drug13" in session:
            drug14 = session["drug13"]
            quantity14 = session["quantity13"]
            size14 = session["size13"]
            species14 = session["species13"]
            for x in productdict.keys():
                if x == drug14:
                    product14 = productdict.get(x)
                    break
                else:
                    continue
        if "drug14" in session:
            drug15 = session["drug14"]
            quantity15 = session["quantity14"]
            size15 = session["size14"]
            species15 = session["species14"]
            for x in productdict.keys():
                if x == drug15:
                    product15 = productdict.get(x)
                    break
                else:
                    continue
        if "drug15" in session:
            drug16 = session["drug15"]
            quantity16 = session["quantity15"]
            size16 = session["size15"]
            species16 = session["species15"]
            for x in productdict.keys():
                if x == drug16:
                    product16 = productdict.get(x)
                    break
                else:
                    continue
        if "drug16" in session:
            drug17 = session["drug16"]
            quantity17 = session["quantity16"]
            size17 = session["size16"]
            species17 = session["species16"]
            for x in productdict.keys():
                if x == drug17:
                    product17 = productdict.get(x)
                    break
                else:
                    continue
        if "drug17" in session:
            drug18 = session["drug17"]
            quantity18 = session["quantity17"]
            size18 = session["size17"]
            species18 = session["species17"]
            for x in productdict.keys():
                if x == drug18:
                    product18 = productdict.get(x)
                    break
                else:
                    continue
        if "drug18" in session:
            drug19 = session["drug18"]
            quantity19 = session["quantity18"]
            size19 = session["size18"]
            species19 = session["species18"]
            for x in productdict.keys():
                if x == drug19:
                    product19 = productdict.get(x)
                    break
                else:
                    continue
        if "drug19" in session:
            drug20 = session["drug19"]
            quantity20 = session["quantity19"]
            size20 = session["size19"]
            species20 = session["species19"]
            for x in productdict.keys():
                if x == drug20:
                    product20 = productdict.get(x)
                    break
                else:
                    continue
        if "drug20" in session:
            drug21 = session["drug20"]
            quantity21 = session["quantity20"]
            size21 = session["size20"]
            species21 = session["species20"]
            for x in productdict.keys():
                if x == drug21:
                    product21 = productdict.get(x)
                    break
                else:
                    continue

        filename = "/home/kyzerfox/mysite/script.xlsx"
        wb = load_workbook("/home/kyzerfox/mysite/Template.xlsx")
        ws = wb["Sheet1"]


        ws["E6"].value = vet
        ws["B5"].value = feedlot
        ws["B44"].value = vet


        iterList = ["A", "B", "C", "D", "E"]#, "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V"]
        varList = [species1, drug1, product1, quantity1, size1, species2, drug2, product2, quantity2, size2, species3, drug3, product3, quantity3, size3, species4, drug4, product4, quantity4, size4, species5, drug5, product5, quantity5, size5, species6, drug6, product6, quantity6, size6, species7, drug7, product7, quantity7, size7, species8, drug8, product8, quantity8, size8, species9, drug9, product9, quantity9, size9, species10, drug10, product10, quantity10, size10, species11, drug11, product11, quantity11, size11, species12, drug12, product12, quantity12, size12, species13, drug13, product13, quantity13, size13, species14, drug14, product14, quantity14, size14, species15, drug15, product15, quantity15, size15, species16, drug16, product16, quantity16, size16, species17, drug17, product17, quantity17, size17, species18, drug18, product18, quantity18, size18, species19, drug19, product19, quantity19, size19, species20, drug20, product20, quantity20, size20, species21, drug21, product21, quantity21, size21]
        varCounter = 0
        for row in range(13, 18):
            for col in iterList:
                ws[col + str(row)].value = varList[varCounter]
                varCounter += 1


        now = datetime.now()
        now2 = now.strftime("%m/%d/%Y")
        ws["E2"].value = now2
        x = datetime.today()
        quarters = ["3/31", "6/30", "9/30", "12/31"]
        quarter = (x.month-1)//3
        returnvar = quarters[quarter] + "/" + str(datetime.today().year)
        ws["E3"].value = returnvar




        vetImgList = ["J. Trent Fox, PhD, DVM", "/home/kyzerfox/mysite/TrentFox.jpg", "Travis McCarty, DVM", "/home/kyzerfox/mysite/TravisMcCarty.png", "Del Miles, DVM, MS", "/home/kyzerfox/mysite/DelMiles.png", "Tom Portillo, DVM", "/home/kyzerfox/mysite/TomPortillo.png", "Randy Hunter, DVM", "/home/kyzerfox/mysite/RandyHunter.png", "Bob Smith, DVM, MS", "", "Miles Theurer, DVM, PhD", "/home/kyzerfox/mysite/MilesTher.png", "Ryn McCollum, DVM", "/home/kyzerfox/mysite/RyanMcCollum.png", "Simon Timmermans, DVM, MS", "/home/kyzerfox/mysite/SimonTimmermans.png", "Jim Sears, DVM", "John Lynch, DVM"]

        count1 = 0
        for i in vetImgList:
            if i == vet:
                count1 += 1
                img = Image(vetImgList[count1])
                ws.add_image(img, "A39")
                break
            else:
                count1 += 1
                continue



        vetlist = ["J. Trent Fox, PhD, DVM", ["NM lic: DVM-2603", "TX lic: 11825", "OK lic: 5353", "CO lic: VET.0009136", "ID lic: V-4034", "AZ lic: 6015", "NV lic: 2342", "NE lic: 3743", "WA lic: VT60551340", "CA lic: vet21235", "KS lic: 7690"], "Travis McCarty, DVM", ["KS lic: 7453", "OK lic: 5127", "CO lic: 0010799", "NE lic: 3860", "WY lic: 1901"], "Del Miles, DVM, MS", ["NE lic: 2648"], "Tom Portillo, DVM", ["TX lic: 8858", "NE lic: 2995", "KS lic: 6650"], "Randy Hunter, DVM", ["NE lic: 2797"], "Bob Smith, DVM, MS", ["TX lic: 4893", "KS lic: 4174"], "Miles Theurer, DVM, PhD", ["KS lic: 8576"], "Ryan McCollum, DVM", ["TX lic: 14141", "KS lic: 8710"], "Simon Timmermans, DVM, MS", ["NE lic: 3530"], "Jim Sears, DVM", "John Lynch, DVM"]

        state = feedlot[-8:-6]
        count2 = 0
        for i in vetlist:
            if i == vet:
                count2 += 1
                for e in vetlist[count2]:
                    if e[0:2] == state:
                        ws["E10"].value = e
                        break
                    else:
                        continue

            else:
                count2 += 1
                continue
        wb.save("Script.xlsx")
        #wb.save(filename)
        wb.close()




        EMAIL_ADDRESS = "office@vrcsllc.com"
        EMAIL_PASSWORD = "gdeqmhqehuphofvy"

        msg = EmailMessage()
        msg["Subject"] = "Authorization Script"
        msg["From"] = EMAIL_ADDRESS
        msg["To"] = "mandy@vrcsllc.com"

        files = ["Script.xlsx"]
        #files = [filename]

        for file in files:
            with open(file, "rb") as f:
                file_data = f.read()
                file_name = f.name
                msg.add_attachment(file_data, maintype="application", subtype="octet_stream", filename=file_name)


        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)

            smtp.send_message(msg)

        return redirect(url_for("end"))
    else:
        return render_template("info.html")

@app.route("/end", methods=["POST", "GET"])
def end():
    return render_template("end.html")


if __name__ == "__main__":
    app.run(debug=True)
